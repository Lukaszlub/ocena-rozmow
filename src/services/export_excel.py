from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List

from src.core.models import EvaluationResult


def export_to_excel(rows: List[EvaluationResult], out_path: Path) -> None:
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "oceny"

    headers = [
        "Imie",
        "Nazwisko",
        "Nazwa pliku",
        "Czas oceny",
        "Wynik %",
        "Gwiazdki",
        "Slowo obrazliwe",
        "Cytat wulgaryzmu",
        "Transkrypcja",
    ]
    ws.append(headers)

    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for r in rows:
        profanity_text = "TAK" if r.profanity_flag else "NIE"
        ws.append(
            [
                r.first_name,
                r.last_name,
                r.file_name,
                r.evaluation_timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                round(r.score_total * 100, 2),
                r.stars,
                profanity_text,
                r.profanity_excerpt,
                r.transcript,
            ]
        )
        cell = ws.cell(row=ws.max_row, column=7)
        cell.fill = red if r.profanity_flag else green

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def default_report_path(reports_dir: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return reports_dir / f"report_{ts}.xlsx"
